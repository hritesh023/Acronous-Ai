import sys
import logging
import subprocess
import tempfile
import os
import re
import httpx
from typing import Optional

logger = logging.getLogger(__name__)

SEARCH_KEYWORDS = [
    "latest", "news", "current", "today", "now", "recent", "update",
    "weather", "forecast", "temperature",
    "stock", "price", "market", "crypto", "bitcoin",
    "score", "winner", "result", "election", "poll",
    "who is", "who won", "what happened",
    "search", "find", "look up", "google",
    "president", "ceo", "leader",
    "2024", "2025", "2026",
    "release", "announce", "launch",
    "trending", "popular", "newest", "top",
]


def needs_web_search(query: str) -> bool:
    ql = query.lower()
    for kw in SEARCH_KEYWORDS:
        if kw in ql:
            return True
    return False


async def web_search(query: str, max_results: int = 5) -> list[dict]:
    try:
        from ddgs import DDGS
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
            return [
                {
                    "title": r.get("title", ""),
                    "link": r.get("href", ""),
                    "snippet": r.get("body", ""),
                }
                for r in results
                if r.get("body")
            ]
    except Exception as e:
        logger.warning(f"DDGS search failed, trying duckduckgo_search: {e}")
        try:
            from duckduckgo_search import DDGS
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=max_results))
                return [
                    {
                        "title": r.get("title", ""),
                        "link": r.get("link", "") or r.get("href", ""),
                        "snippet": r.get("body", "") or r.get("snippet", ""),
                    }
                    for r in results
                    if r.get("body") or r.get("snippet")
                ]
        except Exception as e2:
            logger.warning(f"Both search libraries failed, falling back to HTTP: {e2}")
            pass

        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.post(
                    "https://html.duckduckgo.com/html/",
                    data={"q": query},
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    },
                )
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(resp.text, "lxml")
                results = []
                for item in soup.select(".result")[:max_results]:
                    title_el = item.select_one(".result__title a")
                    snippet_el = item.select_one(".result__snippet")
                    if title_el:
                        results.append({
                            "title": title_el.get_text(strip=True),
                            "link": title_el.get("href", ""),
                            "snippet": snippet_el.get_text(strip=True) if snippet_el else "",
                        })
                return results
        except Exception as e3:
            logger.error(f"All search methods failed: {e3}")
            return []


def format_search_results(results: list[dict]) -> str:
    if not results:
        return "No search results found."
    lines = ["Web search results:", ""]
    for i, r in enumerate(results, 1):
        snippet = r.get("snippet", "")
        title = r.get("title", "")
        link = r.get("link", "")
        lines.append(f"{i}. {title}")
        if snippet:
            lines.append(f"   {snippet}")
        if link:
            lines.append(f"   Source: {link}")
        lines.append("")
    return "\n".join(lines)


async def execute_code(code: str, language: str = "python", timeout: int = 15) -> dict:
    if language not in ("python", "py", "python3"):
        return {"output": "", "error": f"Language '{language}' is not supported yet. Python only for now."}

    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, "script.py")

        wrapped = (
            "import sys, json, math, random, datetime, os, re, collections, itertools, statistics\n"
            "try:\n"
        )
        for line in code.split("\n"):
            wrapped += f"    {line}\n"
        wrapped += "except Exception as e:\n"
        wrapped += "    print(f'ERROR: {e}', file=sys.stderr)\n"

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(wrapped)

        try:
            proc = subprocess.run(
                [sys.executable or "python", filepath],
                capture_output=True,
                text=True,
                timeout=timeout,
                cwd=tmpdir,
                env={**os.environ, "PYTHONIOENCODING": "utf-8"},
            )
            return {
                "output": proc.stdout.strip(),
                "error": proc.stderr.strip() if proc.stderr.strip() else None,
                "exit_code": proc.returncode,
            }
        except subprocess.TimeoutExpired:
            return {"output": "", "error": f"Execution timed out after {timeout} seconds.", "exit_code": -1}
        except Exception as e:
            return {"output": "", "error": str(e), "exit_code": -1}


def format_code_result(result: dict) -> str:
    parts = []
    if result.get("output"):
        parts.append(f"Output:\n```\n{result['output']}\n```")
    if result.get("error"):
        parts.append(f"Error:\n```\n{result['error']}\n```")
    return "\n\n".join(parts) if parts else "No output."


async def process_document(file_bytes: bytes, filename: str) -> dict:
    ext = os.path.splitext(filename)[1].lower()
    text = ""

    try:
        if ext == ".pdf":
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for page in doc:
                text += page.get_text() + "\n"
            doc.close()
        elif ext in (".docx", ".doc"):
            import docx
            doc = docx.Document(file_bytes)
            text = "\n".join(p.text for p in doc.paragraphs)
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"--- Sheet: {sheet_name} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
                text += "\n"
            wb.close()
        elif ext == ".txt":
            text = file_bytes.decode("utf-8", errors="replace")
        elif ext in (".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css", ".json", ".xml", ".md", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".env", ".sh", ".bat", ".ps1"):
            text = file_bytes.decode("utf-8", errors="replace")
        else:
            return {
                "text": f"Unsupported file type: {ext}",
                "type": "error",
                "filename": filename,
            }
    except Exception as e:
        logger.error(f"Document processing error: {e}")
        return {
            "text": f"Error processing {filename}: {e}",
            "type": "error",
            "filename": filename,
        }

    truncated = text[:50000] if len(text) > 50000 else text
    return {
        "text": truncated,
        "type": "document",
        "filename": filename,
        "total_chars": len(text),
    }
